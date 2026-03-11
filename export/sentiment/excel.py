"""
Export Sentiment Analysis to Excel.
"""

from datetime import datetime
from typing import List, Dict

from .utils import _translate_label, _get_sentiment_color

def export_sentiment_to_excel(
    file_path: str,
    results: List[Dict],
    model_type: str = "BERT",
    include_stats: bool = True
) -> bool:
    """
    Export sentiment analysis results to Excel (.xlsx) format.
    
    Args:
        file_path: Target file path
        results: List of sentiment result dictionaries
        model_type: "BERT" or "Online AI"
        include_stats: Include summary statistics sheet
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Prepare data
        data = []
        for r in results:
            data.append({
                'Belge Adı': r.get('title', 'Bilinmeyen'),
                'Duygu Etiketi': _translate_label(r.get('label', 'neutral')),
                'Güven Skoru': f"{r.get('score', 0.5):.1%}",
                'Düz Skor': r.get('score', 0.5),
                'Seviye': r.get('level', 3),
                'Özet': r.get('summary', ''),
                'Model': model_type,
                'Tarih': datetime.now().strftime('%d.%m.%Y %H:%M')
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Duygu Analizi"
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row[key]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color coding for sentiment labels
                if key == 'Duygu Etiketi':
                    color = _get_sentiment_color(row['Duygu Etiketi'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add statistics sheet if requested
        if include_stats:
            _add_stats_sheet(wb, results, model_type)
        
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"Excel export error: {e}")
        return False

def export_hybrid_sentiment_to_excel(file_path: str, results: List[Dict]) -> bool:
    """
    Export hybrid sentiment comparison to Excel.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = []
        for r in results:
            local = r.get('local', {})
            online = r.get('online', {})
            
            data.append({
                'Belge': r.get('title', 'Bilinmeyen'),
                'BERT Duygu': _translate_label(local.get('label', 'neutral')),
                'BERT Skor': local.get('score', 0.5),
                'AI Duygu': _translate_label(online.get('label', 'neutral')),
                'AI Skor': online.get('score', 0.5),
                'AI Güven': online.get('confidence', 0.5),
                'Uyum': 'Evet' if local.get('label') == online.get('label') else 'Hayır',
                'BERT Özet': local.get('summary', ''),
                'AI Özet': online.get('summary', '')
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel with conditional formatting for matches
        wb = Workbook()
        ws = wb.active
        ws.title = "Model Karşılaştırması"
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row[key]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color code matches
                if key == 'Uyum':
                    if row[key] == 'Evet':
                        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"Hybrid export error: {e}")
        return False

def _add_stats_sheet(wb, results, model_type):
    """Add statistics summary sheet to Excel workbook."""
    from openpyxl.styles import Font, PatternFill
    
    ws = wb.create_sheet(title="İstatistikler")
    
    # Calculate stats
    labels = [r.get('label', 'neutral') for r in results]
    label_counts = {}
    for l in labels:
        label_counts[l] = label_counts.get(l, 0) + 1
    
    # Write stats
    ws.cell(row=1, column=1, value="İstatistik").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Değer").font = Font(bold=True)
    
    row = 2
    ws.cell(row=row, column=1, value="Toplam Belge")
    ws.cell(row=row, column=2, value=len(results))
    row += 1
    
    ws.cell(row=row, column=1, value="Model")
    ws.cell(row=row, column=2, value=model_type)
    row += 2
    
    ws.cell(row=row, column=1, value="Duygu Dağılımı").font = Font(bold=True)
    row += 1
    
    for label, count in sorted(label_counts.items()):
        ws.cell(row=row, column=1, value=_translate_label(label))
        ws.cell(row=row, column=2, value=count)
        row += 1
