"""
Excel Processor for LexiScholar
Extracts text from Excel spreadsheets (XLS, XLSX).
"""

from dataclasses import dataclass
from typing import List, Optional
import os

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


@dataclass
class ExcelExtractionResult:
    """Result of Excel text extraction."""
    full_text: str
    sheets: List[str]
    row_count: int
    success: bool
    error: Optional[str] = None


def extract_text(excel_path: str) -> ExcelExtractionResult:
    """
    Extract text from an Excel file (XLS or XLSX).
    
    Args:
        excel_path: Path to the Excel file
        
    Returns:
        ExcelExtractionResult with full text and metadata
    """
    is_xlsx = excel_path.lower().endswith('.xlsx')
    is_xls = excel_path.lower().endswith('.xls')
    
    # Memory Management: Check file size (limit 50MB)
    MAX_SIZE = 50 * 1024 * 1024
    if os.path.exists(excel_path) and os.path.getsize(excel_path) > MAX_SIZE:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error="Dosya çok büyük (Limit: 50MB). Lütfen daha küçük parçalara bölün."
        )
    
    if is_xlsx:
        return _extract_xlsx(excel_path)
    elif is_xls:
        return _extract_xls(excel_path)
    else:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error="Unsupported file format. Use .xls or .xlsx"
        )


def _extract_xlsx(file_path: str) -> ExcelExtractionResult:
    """Extract text from XLSX file using openpyxl."""
    if openpyxl is None:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error="openpyxl library not installed. Run: pip install openpyxl"
        )
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        all_text = []
        total_rows = 0
        
        for sheet_name in sheets:
            ws = wb[sheet_name]
            all_text.append(f"\n{'='*60}")
            all_text.append(f"📊 SAYFA: {sheet_name}")
            all_text.append(f"{'='*60}\n")
            
            for row in ws.iter_rows():
                row_texts = []
                for cell in row:
                    if cell.value is not None:
                        row_texts.append(str(cell.value))
                
                if row_texts:
                    all_text.append(" | ".join(row_texts))
                    total_rows += 1
        
        wb.close()
        
        return ExcelExtractionResult(
            full_text="\n".join(all_text),
            sheets=sheets,
            row_count=total_rows,
            success=True
        )
        
    except FileNotFoundError:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error="File not found"
        )
    except Exception as e:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error=str(e)
        )


def _extract_xls(file_path: str) -> ExcelExtractionResult:
    """Extract text from XLS file using xlrd."""
    if xlrd is None:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error="xlrd library not installed. Run: pip install xlrd"
        )
    
    try:
        wb = xlrd.open_workbook(file_path)
        sheets = wb.sheet_names()
        all_text = []
        total_rows = 0
        
        for sheet_name in sheets:
            ws = wb.sheet_by_name(sheet_name)
            all_text.append(f"\n{'='*60}")
            all_text.append(f"📊 SAYFA: {sheet_name}")
            all_text.append(f"{'='*60}\n")
            
            for row_idx in range(ws.nrows):
                row_texts = []
                for col_idx in range(ws.ncols):
                    cell_value = ws.cell_value(row_idx, col_idx)
                    if cell_value:
                        row_texts.append(str(cell_value))
                
                if row_texts:
                    all_text.append(" | ".join(row_texts))
                    total_rows += 1
        
        return ExcelExtractionResult(
            full_text="\n".join(all_text),
            sheets=sheets,
            row_count=total_rows,
            success=True
        )
        
    except FileNotFoundError:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error="File not found"
        )
    except Exception as e:
        return ExcelExtractionResult(
            full_text="",
            sheets=[],
            row_count=0,
            success=False,
            error=str(e)
        )


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        result = extract_text(sys.argv[1])
        if result.success:
            print(f"Sheets: {result.sheets}")
            print(f"Total rows: {result.row_count}")
            print(f"Total characters: {len(result.full_text)}")
            print("\n--- First 1000 characters ---")
            print(result.full_text[:1000])
        else:
            print(f"Error: {result.error}")
    else:
        print("Usage: python excel_processor.py <excel_file>")
