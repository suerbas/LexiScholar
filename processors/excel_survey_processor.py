"""
Excel Survey Processor for LexiScholar
Extracts structured survey data from Excel spreadsheets (XLS, XLSX).
Supports variable columns and open-ended text columns.
"""

from dataclasses import dataclass
from typing import List, Dict, Optional, Any
import os

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


@dataclass
class SurveyRow:
    """Represents a single row (participant) in a survey."""
    doc_name: str
    variables: Dict[str, Any]  # name -> value
    var_types: Dict[str, str]  # name -> "Tamsayı", "Metin" etc.
    coded_texts: Dict[str, str] # name -> text
    org_coded_names: Dict[str, str] # short_name -> original_long_name (For descriptions)
    row_index: int
    group_name: str = "" # Document group name if selected


@dataclass
class SurveySheetInfo:
    """Metadata about a survey sheet."""
    name: str
    headers: List[str]
    row_count: int


def get_survey_info(excel_path: str) -> List[SurveySheetInfo]:
    """
    Read the sheets and their first row (headers) from an Excel file.
    """
    is_xlsx = excel_path.lower().endswith('.xlsx')
    is_xls = excel_path.lower().endswith('.xls')
    
    if is_xlsx and openpyxl:
        return _get_xlsx_info(excel_path)
    elif is_xls and xlrd:
        return _get_xls_info(excel_path)
    else:
        raise ValueError("Desteklenmeyen dosya formatı veya kütüphane eksik.")


def parse_survey_data(excel_path: str, sheet_name: str, config: Dict[int, Dict[str, str]]) -> List[SurveyRow]:
    """
    Parse survey data based on user configuration.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of the sheet to parse
        config: Mapping of column index (0-based) to configuration:
                {
                    0: {"type": "DOC_NAME", "name": "...", "org_name": "..."},
                    ...
                }
    """
    is_xlsx = excel_path.lower().endswith('.xlsx')
    is_xls = excel_path.lower().endswith('.xls')
    
    if is_xlsx and openpyxl:
        return _parse_xlsx_data(excel_path, sheet_name, config)
    elif is_xls and xlrd:
        return _parse_xls_data(excel_path, sheet_name, config)
    else:
        raise ValueError("Desteklenmeyen dosya formatı veya kütüphane eksik.")


# --- XLSX Implementations (openpyxl) ---

def _get_xlsx_info(file_path: str) -> List[SurveySheetInfo]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    info_list = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        # Get first row
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column) if cell.column else ""
                val = str(cell.value).strip() if (cell is not None and cell.value is not None) else f"Sütun {col_letter}"
                if val:
                    headers.append(val)
            break # only first row
            
        if headers:
            try:
                row_count = ws.max_row
                if row_count is None:
                    # In read_only mode, max_row can be None for some streams.
                    # Fallback to an estimate or 999
                    row_count = 999 
            except Exception:
                row_count = 0
                
            info_list.append(SurveySheetInfo(sheet_name, headers, row_count))
            
    wb.close()
    return info_list


def _parse_xlsx_data(file_path: str, sheet_name: str, config: Dict[int, Dict[str, str]]) -> List[SurveyRow]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sayfa bulunamadı: {sheet_name}")
        
    ws = wb[sheet_name]
    parsed_rows = []
    
    # Skip first row (headers)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Improved default document name fallback
        doc_name = f"{sheet_name} - Satır {row_idx}"
        group_name = ""
        variables = {}
        var_types = {}
        coded_texts = {}
        org_coded_names = {}
        has_data = False
        
        for col_idx, cell in enumerate(row):
            if col_idx not in config:
                continue
                
            conf = config[col_idx]
            ctype = conf.get("type", "IGNORE")
            if ctype == "IGNORE":
                continue
                
            val = str(cell.value).strip() if cell.value is not None else ""
            if not val:
                continue
                
            has_data = True
            
            if conf.get("is_doc_name") or ctype == "DOC_NAME":
                doc_name = val
            if conf.get("is_group_name") or ctype == "GROUP_NAME":
                group_name = val
                
            if ctype == "VARIABLE":
                var_name = conf.get("name", f"Var_{col_idx}")
                variables[var_name] = val
                var_types[var_name] = conf.get("var_type", "Text")
            elif ctype == "CODED_TEXT":
                code_name = conf.get("name", f"Code_{col_idx}")
                org_name = conf.get("org_name", code_name)
                coded_texts[code_name] = val
                org_coded_names[code_name] = org_name
            elif ctype == "BOTH":
                # Hem metin belgesi hem değişken olarak kaydet
                var_name = conf.get("name", f"Var_{col_idx}")
                variables[var_name] = val
                var_types[var_name] = conf.get("var_type", "Metin")
                code_name = conf.get("name", f"Code_{col_idx}")
                org_name = conf.get("org_name", code_name)
                coded_texts[code_name] = val
                org_coded_names[code_name] = org_name
                
        if has_data:
            parsed_rows.append(SurveyRow(
                doc_name=doc_name,
                variables=variables,
                var_types=var_types,
                coded_texts=coded_texts,
                org_coded_names=org_coded_names,
                row_index=row_idx,
                group_name=group_name
            ))
            
    wb.close()
    return parsed_rows


# --- XLS Implementations (xlrd) ---

def _get_xls_info(file_path: str) -> List[SurveySheetInfo]:
    wb = xlrd.open_workbook(file_path)
    info_list = []
    
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows == 0:
            continue
            
        headers = []
        for col_idx in range(ws.ncols):
            val = str(ws.cell_value(0, col_idx)).strip()
            if not val:
                val = f"Sütun {col_idx + 1}"
            headers.append(val)
            
        info_list.append(SurveySheetInfo(sheet_name, headers, ws.nrows))
        
    return info_list


def _parse_xls_data(file_path: str, sheet_name: str, config: Dict[int, Dict[str, str]]) -> List[SurveyRow]:
    wb = xlrd.open_workbook(file_path)
    if sheet_name not in wb.sheet_names():
        raise ValueError(f"Sayfa bulunamadı: {sheet_name}")
        
    ws = wb.sheet_by_name(sheet_name)
    parsed_rows = []
    
    # Skip first row (headers)
    for row_idx in range(1, ws.nrows):
        doc_name = f"{sheet_name} - Satır {row_idx + 1}"
        group_name = ""
        variables = {}
        var_types = {}
        coded_texts = {}
        org_coded_names = {}
        has_data = False
        
        for col_idx in range(ws.ncols):
            if col_idx not in config:
                continue
                
            conf = config[col_idx]
            ctype = conf.get("type", "IGNORE")
            if ctype == "IGNORE":
                continue
                
            val = str(ws.cell_value(row_idx, col_idx)).strip()
            if not val:
                continue
                
            has_data = True
            
            if conf.get("is_doc_name") or ctype == "DOC_NAME":
                doc_name = val
            if conf.get("is_group_name") or ctype == "GROUP_NAME":
                group_name = val
                
            if ctype == "VARIABLE":
                var_name = conf.get("name", f"Var_{col_idx}")
                variables[var_name] = val
                var_types[var_name] = conf.get("var_type", "Text")
            elif ctype == "CODED_TEXT":
                code_name = conf.get("name", f"Code_{col_idx}")
                org_name = conf.get("org_name", code_name)
                coded_texts[code_name] = val
                org_coded_names[code_name] = org_name
            elif ctype == "BOTH":
                # Hem metin belgesi hem değişken olarak kaydet
                var_name = conf.get("name", f"Var_{col_idx}")
                variables[var_name] = val
                var_types[var_name] = conf.get("var_type", "Metin")
                code_name = conf.get("name", f"Code_{col_idx}")
                org_name = conf.get("org_name", code_name)
                coded_texts[code_name] = val
                org_coded_names[code_name] = org_name
                
        if has_data:
            parsed_rows.append(SurveyRow(
                doc_name=doc_name,
                variables=variables,
                var_types=var_types,
                coded_texts=coded_texts,
                org_coded_names=org_coded_names,
                row_index=row_idx + 1,
                group_name=group_name
            ))
            
    return parsed_rows
