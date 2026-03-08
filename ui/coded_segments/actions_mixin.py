"""
Actions and Mixins for Coded Segments Dialog
Handles Export, Word Cloud, and Paraphrase editing.
"""

from PyQt6.QtWidgets import QFileDialog, QMessageBox, QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTextEdit
from PyQt6.QtCore import Qt, QUrl
import os
from ..styles import COLORS, CONTEXT_MENU_STYLE
from ..common_ui import show_info, show_warning, show_error, ask_confirmation

class CodedSegmentsActionsMixin:
    """Handles export, word cloud, and paraphrase operations."""

    def _on_table_context_menu(self, position):
        from PyQt6.QtWidgets import QMenu
        row = self.table.rowAt(position.y())
        if row < 0: return
        self.table.selectRow(row)
        item = self.table.item(row, 0)
        if not item: return
        seg = item.data(Qt.ItemDataRole.UserRole)
        if not seg: return

        menu = QMenu(self)
        menu.setStyleSheet(CONTEXT_MENU_STYLE)
        paraphrase_action = menu.addAction(
            "✍️ Parafraze Ekle/Düzenle…" if not seg.get("paraphrase") else "✍️ Parafrazeyi Düzenle…"
        )
        clear_action = None
        if seg.get("paraphrase"):
            clear_action = menu.addAction("🗑 Parafrazeyi Temizle")

        action = menu.exec(self.table.viewport().mapToGlobal(position))
        if action == paraphrase_action:
            self._edit_paraphrase(seg, row)
        elif clear_action and action == clear_action:
            self._save_paraphrase(seg, "", row)

    def _edit_paraphrase(self, seg: dict, row: int):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTextEdit
        dlg = QDialog(self)
        dlg.setWindowTitle("✍️ Parafraze Ekle / Düzenle")
        dlg.setMinimumSize(520, 280)
        dlg.setModal(True)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(16, 14, 16, 12)
        
        orig = (seg.get("segment_text") or "").strip()
        hint = QLabel(f"<b>Orijinal Metin:</b> {orig[:120]}{'…' if len(orig) > 120 else ''}")
        hint.setWordWrap(True)
        hint.setStyleSheet(f"font-size:11px; color:{COLORS.get('text_secondary','#64748B')}; background:{COLORS.get('primary_50','#EEF2FF')}; border-left:3px solid #4F46E5; padding:6px 10px; border-radius:4px;")
        lay.addWidget(hint)
        
        lbl = QLabel("Parafrazeni yaz (araştırmacının kendi sözcükleriyle özeti):")
        lbl.setStyleSheet("font-size:12px; font-weight:600; color:#1E293B;")
        lay.addWidget(lbl)
        
        editor = QTextEdit()
        editor.setPlainText(seg.get("paraphrase") or "")
        editor.setStyleSheet("QTextEdit { border:1.5px solid #CBD5E1; border-radius:8px; padding:8px 10px; font-size:13px; color:#1E293B; background:#FFFFFF;} QTextEdit:focus { border-color:#4F46E5; }")
        editor.setFixedHeight(100)
        lay.addWidget(editor)
        
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("İptal")
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_cancel)
        btn_save = QPushButton("✅ Kaydet")
        btn_save.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_save)
        lay.addLayout(btn_row)
        
        if dlg.exec():
            self._save_paraphrase(seg, editor.toPlainText(), row)

    def _save_paraphrase(self, seg: dict, text: str, row: int):
        seg_id = seg.get("id")
        if not seg_id or not self._segment_dao:
            show_warning(self, "Hata", "Segment DAO bağlı değil.")
            return
        
        old_text = seg.get("paraphrase") or ""
        if self._command_stack is not None:
            from ..commands import UpdateParaphraseCommand
            cmd = UpdateParaphraseCommand(self._segment_dao, seg_id, text, old_text)
            self._command_stack.push(cmd)
            ok = True
        else:
            ok = self._segment_dao.update_paraphrase(seg_id, text)
        
        if ok:
            seg["paraphrase"] = text
            para_item = self.table.item(row, self.COL_PARAPHRASE)
            if para_item:
                para_item.setText((text[:80] + "…") if len(text) > 80 else text)
            self._update_preview(seg)
        else:
            show_warning(self, "Hata", "Parafraze kaydedilemedi.")

    def _export_to_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Excel'e Aktar", f"KodlanmisSegmentler_{self._code_name.replace(' ', '_')}.xlsx", "Excel Dosyası (*.xlsx);;CSV Dosyası (*.csv)")
        if not file_path: return
        try:
            if file_path.endswith(".csv"): self._export_csv(file_path)
            else: self._export_xlsx(file_path)
            show_info(self, "Başarılı", f"Veriler dışa aktarıldı:\n{file_path}")
        except Exception as e:
            show_error(self, "Hata", f"Dışa aktarma başarısız:\n{e}")

    def _export_xlsx(self, file_path: str):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from .base import COLUMNS
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Kodlanmış - {self._code_name[:28]}"
        col_names = [c[0] for c in COLUMNS]
        col_names[-1] = "Tam Metin"
        header_fill = PatternFill("solid", fgColor="4F46E5")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style='thin', color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill, cell.font, cell.border = header_fill, header_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for ri, seg in enumerate(self._segments, start=2):
            weight = seg.get("weight", 0)
            row_data = [seg.get("comment", ""), seg.get("folder_name", "—"), seg.get("document_title", "—"), seg.get("code_name", "") or self._code_name, seg.get("start_pos", 0), seg.get("end_pos", 0), f"{weight} ⭐" if weight else "", seg.get("paraphrase", ""), (seg.get("segment_text", "") or "").strip()]
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=(ci == len(row_data)))
                cell.border = border
                if ri % 2 == 0: cell.fill = PatternFill("solid", fgColor="F8FAFC")
        col_widths = [16, 17, 22, 16, 12, 12, 10, 30, 60]
        for ci, w in enumerate(col_widths, start=1): ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        wb.save(file_path)

    def _export_csv(self, file_path: str):
        import csv
        from .base import COLUMNS
        col_names = [c[0] for c in COLUMNS]
        col_names[-1] = "Tam Metin"
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(col_names)
            for seg in self._segments:
                weight = seg.get("weight", 0)
                writer.writerow([seg.get("comment", ""), seg.get("folder_name", "—"), seg.get("document_title", "—"), seg.get("code_name", "") or self._code_name, seg.get("start_pos", 0), seg.get("end_pos", 0), f"{weight} ★" if weight else "", seg.get("paraphrase", ""), (seg.get("segment_text", "") or "").strip()])

    def _show_word_cloud(self):
        from collections import Counter
        import re
        from visualizations.word_cloud import generate_word_cloud_html
        from analysis.analysis_tools import STOP_WORDS
        counter = Counter()
        pattern = re.compile(r'\b[a-zçğıöşü]{3,}\b')
        for seg in self._segments:
            text = (seg.get("segment_text", "") or "").lower()
            text = re.sub(r'<[^>]+>', ' ', text)
            words = pattern.findall(text)
            counter.update([w for w in words if w not in STOP_WORDS])
        word_freq = counter.most_common(150)
        if not word_freq: return
        file_path = generate_word_cloud_html(word_freq)
        self._open_word_cloud_dialog(file_path)

    def _open_word_cloud_dialog(self, html_path: str):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QWidget
        from PyQt6.QtCore import QUrl
        dlg = QDialog(self)
        dlg.setWindowTitle(f"☁️ Kelime Bulutu — {self._code_name}")
        dlg.setMinimumSize(900, 600)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(0, 0, 0, 0)
        try:
            from PyQt6.QtWebEngineWidgets import QWebEngineView
            browser = QWebEngineView()
            browser.setUrl(QUrl.fromLocalFile(html_path))
            layout.addWidget(browser)
        except ImportError:
            import webbrowser
            webbrowser.open(f'file://{html_path}')
            return
        btn_container = QWidget()
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.addStretch()
        close_btn = QPushButton("✕ Kapat")
        close_btn.clicked.connect(dlg.accept)
        btn_layout.addWidget(close_btn)
        layout.addWidget(btn_container)
        dlg.exec()
