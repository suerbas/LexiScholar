"""
Coded Segments Analysis Dialog (Consolidated)
Combines Base, Table Manager, and Actions into a single file to prevent import/mixin crashes.
"""

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QSplitter,
    QTextEdit, QLabel, QWidget, QPushButton,
    QTableWidget, QTableWidgetItem, QAbstractItemView, QHeaderView,
    QMenu, QFileDialog, QMessageBox, QFrame
)
from PyQt6.QtCore import Qt, pyqtSignal, QSize, QUrl
from PyQt6.QtGui import QColor
import os
from ..common_ui import show_info, show_warning, show_error, ask_confirmation

# --- Styles ---
COLORS = {
    'bg_main': '#F1F5F9',
    'bg_panel': '#FFFFFF',
    'bg_hover': '#F8FAFC',
    'bg_selected': '#E0E7FF',
    'primary_50': '#EEF2FF',
    'primary_100': '#E0E7FF',
    'primary_600': '#4F46E5',
    'text_primary': '#1E293B',
    'text_secondary': '#475569',
    'text_muted': '#94A3B8',
    'border': '#E2E8F0',
    'border_hover': '#CBD5E1',
}

CONTEXT_MENU_STYLE = """
QMenu {
    background-color: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 6px;
    padding: 4px;
}
QMenu::item {
    padding: 6px 24px;
    border-radius: 4px;
    color: #1E293B;
}
QMenu::item:selected {
    background-color: #F1F5F9;
    color: #0F172A;
}
"""

_DIALOG_EXTRA_STYLE = f"""
QDialog {{
    background-color: {COLORS['bg_main']};
    font-family: 'Segoe UI', sans-serif;
}}
QPushButton#toolBtn {{
    background-color: #FFFFFF;
    color: {COLORS['text_secondary']};
    border: 1px solid {COLORS['border']};
    border-radius: 6px;
    padding: 6px 14px;
    font-size: 9pt;
    font-weight: 600;
}}
QPushButton#toolBtn:hover {{
    background-color: {COLORS['bg_hover']};
    border-color: {COLORS['border_hover']};
    color: {COLORS['text_primary']};
}}
QPushButton#primaryBtn {{
    background-color: #4F46E5;
    color: #FFFFFF;
    border: none;
    border-radius: 6px;
    padding: 6px 16px;
    font-size: 9pt;
    font-weight: 700;
}}
QPushButton#primaryBtn:hover {{ background-color: #4338CA; }}
QWidget#headerBar {{
    background-color: {COLORS['bg_panel']};
    border-bottom: 2px solid {COLORS['border_hover']};
}}
QTextEdit#previewEdit {{
    background: transparent;
    color: #1E293B;
    border: none;
    font-size: 11pt;
    line-height: 1.4;
    padding: 14px 20px;
    selection-background-color: #EEF2FF;
}}
QTableWidget {{
    background-color: #FFFFFF;
    gridline-color: {COLORS['border']};
    border: none;
    font-size: 9pt;
    color: {COLORS['text_primary']};
    selection-background-color: {COLORS['bg_selected']};
}}
QHeaderView::section {{
    background-color: {COLORS['primary_100']};
    color: {COLORS['text_secondary']};
    padding: 8px 10px;
    border: none;
    border-right: 1px solid {COLORS['border_hover']};
    border-bottom: 2px solid {COLORS['border_hover']};
    font-weight: 700;
    text-transform: uppercase;
}}
QLabel#statusLabel {{
    color: {COLORS['text_muted']};
    font-size: 8.5pt;
    padding: 4px 0px;
}}
QLabel#countBadge {{
    background-color: #4F46E5;
    color: #FFFFFF;
    border-radius: 10px;
    padding: 2px 10px;
    font-size: 8pt;
    font-weight: 700;
}}
QSplitter::handle:vertical {{
    background-color: {COLORS['border_hover']};
    height: 4px;
    margin: 0 8px;
}}
QSplitter::handle:vertical:hover {{ background-color: #4F46E5; }}
"""

COLUMNS = [
    ("Notlar",         120),
    ("Belge Grubu",    130),
    ("Belge Adı",      160),
    ("Kod Adı",        120),
    ("Başlangıç",       80),
    ("Bitiş",           80),
    ("Ağırlık",         70),
    ("Parafraze",      160),
    ("Önizleme Metni", 280),
]

class CodedSegmentsWidget(QWidget):
    """
    Widget version of Coded Segments for tabbed interface.
    """
    segment_navigate_requested = pyqtSignal(int, int)

    COL_COMMENT  = 0
    COL_GROUP    = 1
    COL_DOC_NAME = 2
    COL_CODE     = 3
    COL_START    = 4
    COL_END      = 5
    COL_WEIGHT   = 6
    COL_PARAPHRASE = 7
    COL_PREVIEW  = 8

    def __init__(self, segments: list, code_name: str = "",
                 code_color: str = "#4F46E5", segment_dao=None,
                 command_stack=None, parent=None):
        super().__init__(parent)
        self._segments = segments
        self._code_name = code_name
        self._code_color = code_color
        self._segment_dao = segment_dao
        self._command_stack = command_stack
        
        self.setStyleSheet(_DIALOG_EXTRA_STYLE)
        self._setup_ui()
        self._populate_table()

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Header
        self.header = self._build_header()
        root.addWidget(self.header)

        # Splitter
        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.setHandleWidth(6)
        
        self.preview_panel = self._build_preview_panel()
        self.table_panel = self._build_table_panel()
        
        splitter.addWidget(self.preview_panel)
        splitter.addWidget(self.table_panel)
        
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 7)
        splitter.setSizes([220, 480])
        
        root.addWidget(splitter, 1)

        # Footer
        self.footer = self._build_footer()
        root.addWidget(self.footer)

    def _build_header(self) -> QWidget:
        bar = QWidget()
        bar.setObjectName("headerBar")
        bar.setFixedHeight(52)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(16, 0, 16, 0)
        layout.setSpacing(12)

        circle = QLabel()
        circle.setFixedSize(14, 14)
        circle.setStyleSheet(f"background-color:{self._code_color}; border-radius:7px;")
        layout.addWidget(circle)

        title = QLabel(
            f"<b style='font-size:12pt;color:{COLORS['text_primary']};'>Kodlanmış Bölümler</b>"
            f"&nbsp;&nbsp;<span style='color:{COLORS['text_muted']};font-size:10pt;'>— {self._code_name}</span>"
        )
        title.setTextFormat(Qt.TextFormat.RichText)
        layout.addWidget(title)
        layout.addStretch()

        self.lbl_count = QLabel("0 segment")
        self.lbl_count.setObjectName("countBadge")
        layout.addWidget(self.lbl_count)

        btn_wordcloud = QPushButton("☁️ Kelime Bulutu")
        btn_wordcloud.setObjectName("toolBtn")
        btn_wordcloud.clicked.connect(self._show_word_cloud)
        layout.addWidget(btn_wordcloud)

        btn_export = QPushButton("📥 Excel'e Aktar")
        btn_export.setObjectName("primaryBtn")
        btn_export.clicked.connect(self._export_to_excel)
        layout.addWidget(btn_export)

        return bar

    def _build_preview_panel(self):
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        strip = QLabel("  📄  SEGMENT ÖNİZLEME")
        strip.setStyleSheet(
            f"background-color:{COLORS['primary_50']}; color:{COLORS['text_secondary']}; "
            f"font-size:8pt; font-weight:700; padding:6px 12px; border-bottom:1px solid {COLORS['border']};"
        )
        layout.addWidget(strip)
        
        self.preview_edit = QTextEdit()
        self.preview_edit.setObjectName("previewEdit")
        self.preview_edit.setReadOnly(True)
        layout.addWidget(self.preview_edit)
        return container

    def _build_table_panel(self):
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.table_strip = QLabel("  📊  KODLANMIŞ BÖLÜMLER TABLOSU")
        self.table_strip.setStyleSheet(
            f"background-color:{COLORS['primary_50']}; color:{COLORS['text_secondary']}; "
            f"font-size:8pt; font-weight:700; padding:6px 12px; border-bottom:1px solid {COLORS['border']};"
        )
        layout.addWidget(self.table_strip)

        self.table = QTableWidget()
        self.table.setColumnCount(len(COLUMNS))
        self.table.setHorizontalHeaderLabels([c[0] for c in COLUMNS])
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.setShowGrid(True)

        hdr = self.table.horizontalHeader()
        for i, (_, width) in enumerate(COLUMNS):
            self.table.setColumnWidth(i, width)
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        self.table.verticalHeader().setDefaultSectionSize(34)

        self.table.itemSelectionChanged.connect(self._on_row_selected)
        self.table.cellDoubleClicked.connect(self._on_row_double_clicked)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_table_context_menu)

        layout.addWidget(self.table)
        return container

    def _build_footer(self):
        footer = QWidget()
        footer.setStyleSheet(f"background-color:{COLORS['bg_panel']}; border-top:1px solid {COLORS['border']};")
        footer.setFixedHeight(40)
        layout = QHBoxLayout(footer)
        layout.setContentsMargins(16, 4, 16, 4)
        
        self.lbl_status = QLabel("Segment seçmek için tabloya tıklayın.")
        self.lbl_status.setObjectName("statusLabel")
        layout.addWidget(self.lbl_status)
        layout.addStretch()
        
        tip = QLabel("💡 Çift tıklayarak belgedeki konuma git")
        tip.setObjectName("statusLabel")
        layout.addWidget(tip)
        return footer

    # --- Actions ---

    def _on_row_selected(self):
        items = self.table.selectedItems()
        if not items: return
        row = items[0].row()
        item = self.table.item(row, 0)
        if item:
            seg = item.data(Qt.ItemDataRole.UserRole)
            self._update_preview(seg)

    def _on_row_double_clicked(self, row, col):
        item = self.table.item(row, 0)
        if item:
            seg = item.data(Qt.ItemDataRole.UserRole)
            if not seg: return
            
            # Double clicking Notes column opens edit note dialog
            if col == self.COL_COMMENT:
                self._edit_comment(seg, row)
            else:
                self.segment_navigate_requested.emit(seg.get('document_id'), seg.get('id'))

    def _on_table_context_menu(self, position):
        row = self.table.rowAt(position.y())
        if row < 0: return
        self.table.selectRow(row)
        item = self.table.item(row, 0)
        if not item: return
        seg = item.data(Qt.ItemDataRole.UserRole)
        if not seg: return

        menu = QMenu(self)
        menu.setStyleSheet(CONTEXT_MENU_STYLE)
        comment_action = menu.addAction(
            "📝 Not Ekle/Düzenle…" if not seg.get("comment") else "📝 Notu Düzenle…"
        )
        paraphrase_action = menu.addAction(
            "✍️ Parafraze Ekle/Düzenle…" if not seg.get("paraphrase") else "✍️ Parafrazeyi Düzenle…"
        )
        menu.addSeparator()
        if seg.get("paraphrase"):
            clear_action = menu.addAction("🗑 Parafrazeyi Temizle")

        action = menu.exec(self.table.viewport().mapToGlobal(position))
        if action == comment_action:
            self._edit_comment(seg, row)
        elif action == paraphrase_action:
            self._edit_paraphrase(seg, row)
        elif clear_action and action == clear_action:
            self._save_paraphrase(seg, "", row)

    def _populate_table(self):
        segs = self._segments
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self.table.setRowCount(len(segs))

        color = QColor(self._code_color)
        color.setAlpha(40)

        for row, seg in enumerate(segs):
            doc_title = seg.get("document_title", "") or "—"
            code_name = seg.get("code_name", "") or self._code_name
            start_pos = seg.get("start_pos", 0)
            end_pos   = seg.get("end_pos", 0)
            weight    = seg.get("weight", 0)
            seg_text  = (seg.get("segment_text", "") or "").strip()
            comment   = seg.get("comment", "") or ""
            paraphrase = seg.get("paraphrase", "") or ""
            folder_name = seg.get("folder_name", "")
            doc_group = folder_name if folder_name else "—"

            preview_short = (seg_text[:120] + "…") if len(seg_text) > 120 else seg_text
            para_short    = (paraphrase[:80] + "…") if len(paraphrase) > 80 else paraphrase

            cells = [comment, doc_group, doc_title, code_name, str(start_pos), str(end_pos), self._weight_stars(weight), para_short, preview_short]

            for col, text in enumerate(cells):
                item = QTableWidgetItem(text)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                if col == self.COL_WEIGHT:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                elif col in (self.COL_START, self.COL_END):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, col, item)

            self.table.item(row, 0).setData(Qt.ItemDataRole.UserRole, seg)
            doc_item = self.table.item(row, self.COL_DOC_NAME)
            if doc_item:
                doc_item.setBackground(color)

        self.table.setSortingEnabled(True)
        count = len(segs)
        self.lbl_count.setText(f"{count} segment" if count != 1 else "1 segment")
        self.table_strip.setText(f"  📊  KODLANMIŞ BÖLÜMLER TABLOSU — {count} kayıt")

    def _update_preview(self, seg: dict):
        full_text = (seg.get("segment_text", "") or "").strip()
        doc_title  = seg.get("document_title", "") or "Belge"
        code_name  = seg.get("code_name", "") or self._code_name
        weight     = seg.get("weight", 0)
        start_pos  = seg.get("start_pos", 0)
        end_pos    = seg.get("end_pos", 0)
        color      = self._code_color
        paraphrase = (seg.get("paraphrase") or "").strip()

        para_html = ""
        if paraphrase:
            para_html = (
                f"<div style='margin:10px 0 8px 0; padding:8px 14px;"
                f" background:#FFFBEB; border-left:4px solid #F59E0B;"
                f" border-radius:5px;'>"
                f"<span style='font-size:8pt; font-weight:700; color:#92400E; "
                f"letter-spacing:0.5px;'>✍️ PARAFRAZE</span><br>"
                f"<span style='font-size:10pt; color:#1E293B;'>{paraphrase}</span>"
                f"</div>"
            )

        html = (
            f"<div style='margin-bottom:10px;'>"
            f"<span style='display:inline-block; background:{color}22; "
            f"border-left:4px solid {color}; padding:3px 8px; border-radius:3px;"
            f"font-size:9pt; font-weight:700; color:{color};'>"
            f"🏷️ {code_name}</span>"
            f"&nbsp;&nbsp;"
            f"<span style='color:{COLORS['text_muted']};font-size:9pt;'>"
            f"📄 {doc_title} &nbsp;|&nbsp; "
            f"Pos {start_pos}–{end_pos} &nbsp;|&nbsp; "
            f"{weight} ⭐"
            f"</span>"
            f"</div>"
            f"{para_html}"
            f"<div style='font-size:11pt; line-height:1.4; color:{COLORS['text_primary']}; white-space:pre-wrap;'>"
            f"{full_text}"
            f"</div>"
        )
        self.preview_edit.setHtml(html)
        self.lbl_status.setText(f"📄 {doc_title}  ·  Karakter: {start_pos}–{end_pos}  ·  Ağırlık: {weight} ⭐")

    @staticmethod
    def _weight_stars(weight: int) -> str:
        try:
            w = max(1, min(5, int(weight)))
            return f"{w} ⭐"
        except (TypeError, ValueError):
            return "—"

    # --- Feature Mixin Methods (Inlined) ---
    def _edit_paraphrase(self, seg: dict, row: int):
        dlg = QDialog(self)
        dlg.setWindowTitle("✍️ Parafraze Ekle / Düzenle")
        dlg.setMinimumSize(520, 280)
        dlg.setModal(True)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(16, 14, 16, 12)
        
        orig = (seg.get("segment_text") or "").strip()
        hint = QLabel(f"<b>Orijinal Metin:</b> {orig[:120]}{'…' if len(orig) > 120 else ''}")
        hint.setWordWrap(True)
        hint.setStyleSheet(f"font-size:11px; color:{COLORS.get('text_secondary','#64748B')}; background:{COLORS.get('primary_50','#EEF2FF')}; border-left:3px solid #4F46E5; padding:6px 10px; border-radius:4px;")
        lay.addWidget(hint)
        
        lbl = QLabel("Parafrazeni yaz (araştırmacının kendi sözcükleriyle özeti):")
        lbl.setStyleSheet("font-size:12px; font-weight:600; color:#1E293B;")
        lay.addWidget(lbl)
        
        editor = QTextEdit()
        editor.setPlainText(seg.get("paraphrase") or "")
        editor.setStyleSheet("QTextEdit { border:1.5px solid #CBD5E1; border-radius:8px; padding:8px 10px; font-size:13px; color:#1E293B; background:#FFFFFF;} QTextEdit:focus { border-color:#4F46E5; }")
        editor.setFixedHeight(100)
        lay.addWidget(editor)
        
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("İptal")
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_cancel)
        btn_save = QPushButton("✅ Kaydet")
        btn_save.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_save)
        lay.addLayout(btn_row)
        
        if dlg.exec():
            self._save_paraphrase(seg, editor.toPlainText(), row)

    def _save_paraphrase(self, seg: dict, text: str, row: int):
        seg_id = seg.get("id")
        if not seg_id or not self._segment_dao:
            show_warning(self, "Hata", "Segment DAO bağlı değil.")
            return
        
        old_text = seg.get("paraphrase") or ""
        if self._command_stack is not None:
            # Assume command class exists elsewhere or import it
            # For simplicity, let's use DAO directly if command stack fails or is complex
            # But we should respect undo. Let's try to import.
            try:
                from ..commands import UpdateParaphraseCommand
                cmd = UpdateParaphraseCommand(self._segment_dao, seg_id, text, old_text)
                self._command_stack.push(cmd)
                ok = True
            except ImportError:
                 ok = self._segment_dao.update_paraphrase(seg_id, text)
        else:
            ok = self._segment_dao.update_paraphrase(seg_id, text)
        
        if ok:
            seg["paraphrase"] = text
            para_item = self.table.item(row, self.COL_PARAPHRASE)
            if para_item:
                para_item.setText((text[:80] + "…") if len(text) > 80 else text)
            self._update_preview(seg)
        else:
            show_warning(self, "Hata", "Parafraze kaydedilemedi.")

    def _export_to_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Excel'e Aktar", f"KodlanmisSegmentler_{self._code_name.replace(' ', '_')}.xlsx", "Excel Dosyası (*.xlsx);;CSV Dosyası (*.csv)")
        if not file_path: return
        try:
            if file_path.endswith(".csv"): self._export_csv(file_path)
            else: self._export_xlsx(file_path)
            show_info(self, "Başarılı", f"Veriler dışa aktarıldı:\n{file_path}")
        except Exception as e:
            show_error(self, "Hata", f"Dışa aktarma başarısız:\n{e}")

    def _export_xlsx(self, file_path: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
             show_warning(self, "Eksik Kütüphane", "openpyxl yüklü değil.")
             return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Kodlanmış - {self._code_name[:28]}"
        col_names = [c[0] for c in COLUMNS]
        col_names[-1] = "Tam Metin"
        header_fill = PatternFill("solid", fgColor="4F46E5")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style='thin', color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill, cell.font, cell.border = header_fill, header_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for ri, seg in enumerate(self._segments, start=2):
            weight = seg.get("weight", 0)
            row_data = [seg.get("comment", ""), seg.get("folder_name", "—"), seg.get("document_title", "—"), seg.get("code_name", "") or self._code_name, seg.get("start_pos", 0), seg.get("end_pos", 0), f"{weight} ⭐" if weight else "", seg.get("paraphrase", ""), (seg.get("segment_text", "") or "").strip()]
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=(ci == len(row_data)))
                cell.border = border
                if ri % 2 == 0: cell.fill = PatternFill("solid", fgColor="F8FAFC")
        col_widths = [16, 17, 22, 16, 12, 12, 10, 30, 60]
        for ci, w in enumerate(col_widths, start=1): ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        wb.save(file_path)

    def _export_csv(self, file_path: str):
        import csv
        col_names = [c[0] for c in COLUMNS]
        col_names[-1] = "Tam Metin"
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(col_names)
            for seg in self._segments:
                weight = seg.get("weight", 0)
                writer.writerow([seg.get("comment", ""), seg.get("folder_name", "—"), seg.get("document_title", "—"), seg.get("code_name", "") or self._code_name, seg.get("start_pos", 0), seg.get("end_pos", 0), f"{weight} ★" if weight else "", seg.get("paraphrase", ""), (seg.get("segment_text", "") or "").strip()])

    def _show_word_cloud(self):
        from collections import Counter
        import re
        from visualizations.word_cloud import generate_word_cloud_html
        from analysis.analysis_tools import STOP_WORDS
        counter = Counter()
        pattern = re.compile(r'\b[a-zçğıöşü]{3,}\b')
        for seg in self._segments:
            text = (seg.get("segment_text", "") or "").lower()
            text = re.sub(r'<[^>]+>', ' ', text)
            words = pattern.findall(text)
            counter.update([w for w in words if w not in STOP_WORDS])
        word_freq = counter.most_common(150)
        if not word_freq: return
        file_path = generate_word_cloud_html(word_freq)
        self._open_word_cloud_dialog(file_path)

    def _open_word_cloud_dialog(self, html_path: str):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"☁️ Kelime Bulutu — {self._code_name}")
        dlg.setMinimumSize(900, 600)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(0, 0, 0, 0)
        try:
            from PyQt6.QtWebEngineWidgets import QWebEngineView
            browser = QWebEngineView()
            browser.setUrl(QUrl.fromLocalFile(html_path))
            layout.addWidget(browser)
        except ImportError:
            import webbrowser
            webbrowser.open(f'file://{html_path}')
            return
        btn_container = QWidget()
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.addStretch()
        close_btn = QPushButton("✕ Kapat")
        close_btn.clicked.connect(dlg.accept)
        btn_layout.addWidget(close_btn)
        layout.addWidget(btn_container)
        dlg.exec()

    def _edit_comment(self, seg: dict, row: int):
        dlg = QDialog(self)
        dlg.setWindowTitle("📝 Not Ekle / Düzenle")
        dlg.setMinimumSize(480, 240)
        dlg.setModal(True)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(16, 14, 16, 12)
        
        lbl = QLabel("Bu segment için notunu (yorumunu) yaz:")
        lbl.setStyleSheet("font-size:12px; font-weight:600; color:#1E293B;")
        lay.addWidget(lbl)
        
        editor = QTextEdit()
        editor.setPlainText(seg.get("comment") or "")
        editor.setStyleSheet("QTextEdit { border:1.5px solid #CBD5E1; border-radius:8px; padding:8px 10px; font-size:13px; color:#1E293B; background:#FFFFFF;} QTextEdit:focus { border-color:#4F46E5; }")
        editor.setFixedHeight(100)
        lay.addWidget(editor)
        
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("İptal")
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_cancel)
        btn_save = QPushButton("✅ Kaydet")
        btn_save.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_save)
        lay.addLayout(btn_row)
        
        if dlg.exec():
            self._save_comment(seg, editor.toPlainText(), row)

    def _save_comment(self, seg: dict, text: str, row: int):
        seg_id = seg.get("id")
        if not seg_id or not self._segment_dao:
            show_warning(self, "Hata", "Segment DAO bağlı değil.")
            return
        
        old_text = seg.get("comment") or ""
        if self._command_stack is not None:
            try:
                from ..commands import UpdateCommentCommand
                cmd = UpdateCommentCommand(self._segment_dao, seg_id, text, old_text)
                self._command_stack.push(cmd)
                ok = True
            except ImportError:
                  ok = self._segment_dao.update_comment(seg_id, text)
        else:
            ok = self._segment_dao.update_comment(seg_id, text)
        
        if ok:
            seg["comment"] = text
            comment_item = self.table.item(row, self.COL_COMMENT)
            if comment_item:
                comment_item.setText(text) # In table, we show full text for comments/notes usually or truncated
            self._update_preview(seg)
        else:
            show_warning(self, "Hata", "Not kaydedilemedi.")


class CodedSegmentsDialog(QDialog):
    """
    Standalone dialog version of retrieved segments.
    Wraps CodedSegmentsWidget for backward compatibility.
    """
    segment_navigate_requested = pyqtSignal(int, int)

    def __init__(self, segments: list, code_name: str = "",
                 code_color: str = "#4F46E5", segment_dao=None,
                 command_stack=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"🔎 Kodlanmış Bölümler — {code_name}")
        self.setMinimumSize(QSize(960, 640))
        self.resize(QSize(1200, 750))
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        self.widget = CodedSegmentsWidget(
            segments, code_name, code_color, segment_dao, command_stack, self
        )
        # Relay signal
        self.widget.segment_navigate_requested.connect(self.segment_navigate_requested.emit)
        
        # Connect close button in widget header to dialog close
        # It's the only toolBtn in the header
        btn_close = self.widget.header.findChild(QPushButton, "toolBtn")
        if btn_close:
            try:
                btn_close.clicked.disconnect()
            except: pass
            btn_close.clicked.connect(self.close)
            
        layout.addWidget(self.widget)
